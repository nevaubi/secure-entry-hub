"""
Excel file updater using openpyxl.

Handles updating specific cells in Excel files based on AI instructions.
"""

from copy import copy
from pathlib import Path
from typing import Any
import openpyxl
from openpyxl.utils import column_index_from_string


class ExcelUpdater:
    """Updates Excel files based on AI instructions."""

    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.changes_made = 0

    def update_cell(self, sheet_name: str, cell_ref: str, value: Any) -> bool:
        try:
            if sheet_name not in self.workbook.sheetnames:
                print(f"Sheet '{sheet_name}' not found")
                return False

            sheet = self.workbook[sheet_name]
            sheet[cell_ref] = value
            self.changes_made += 1

            cell = sheet[cell_ref]
            if cell.column == 2:  # Column B
                source = sheet.cell(row=cell.row, column=3)
                cell.font = copy(source.font)
                cell.fill = copy(source.fill)
                cell.alignment = copy(source.alignment)
                cell.border = copy(source.border)
                cell.number_format = source.number_format

            print(f"Updated {sheet_name}!{cell_ref} = {value}")
            return True

        except Exception as e:
            print(f"Error updating cell {sheet_name}!{cell_ref}: {e}")
            return False

    def update_cells_batch(self, updates: list[dict]) -> int:
        successful = 0
        for update in updates:
            if self.update_cell(update["sheet_name"], update["cell_ref"], update["value"]):
                successful += 1
        return successful

    def save(self) -> bool:
        try:
            self.workbook.save(self.file_path)
            print(f"Saved {self.file_path} with {self.changes_made} changes")
            return True
        except Exception as e:
            print(f"Error saving {self.file_path}: {e}")
            return False

    def close(self):
        self.workbook.close()

    def insert_new_period_column(
        self, sheet_name: str, date_header: str, period_header: str
    ) -> dict:
        try:
            if sheet_name not in self.workbook.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}

            sheet = self.workbook[sheet_name]
            sheet.insert_cols(2)

            sheet.cell(row=1, column=2).value = date_header
            sheet.cell(row=2, column=2).value = period_header
            self.changes_made += 2

            for row in [1, 2]:
                source_cell = sheet.cell(row=row, column=3)
                target_cell = sheet.cell(row=row, column=2)
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.border = copy(source_cell.border)
                target_cell.number_format = source_cell.number_format

            data_rows = []
            row_map = []
            for row_idx in range(3, (sheet.max_row or 2) + 1):
                cell_c = sheet.cell(row=row_idx, column=3)
                if cell_c.value is not None:
                    data_rows.append(row_idx)
                    label = sheet.cell(row=row_idx, column=1).value or ""
                    row_map.append({"row": row_idx, "label": str(label).strip(), "cell": f"B{row_idx}"})

            print(f"Inserted new column B in {sheet_name}: {date_header} / {period_header}")
            print(f"  {len(data_rows)} rows need data")

            cell_list = ", ".join(f"B{r} ({m['label']})" for r, m in zip(data_rows[:50], row_map[:50]))
            if len(data_rows) > 50:
                cell_list += f"... ({len(data_rows)} total)"

            return {
                "success": True,
                "data_rows": data_rows,
                "row_map": row_map,
                "total_rows_needing_data": len(data_rows),
                "message": f"New column B inserted with headers '{date_header}' / '{period_header}'. Fill these cells: {cell_list}",
            }

        except Exception as e:
            print(f"Error inserting column in {sheet_name}: {e}")
            return {"success": False, "error": str(e)}

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


def update_file(file_path: Path, updates: list[dict]) -> tuple[bool, int]:
    with ExcelUpdater(file_path) as updater:
        count = updater.update_cells_batch(updates)
        if count > 0:
            success = updater.save()
            return success, count
        return True, 0
