"""
Excel schema analysis using Claude.

Dynamically understands the structure of Excel files to determine
what data needs to be updated.
"""

import os
from pathlib import Path
from typing import Any
import openpyxl
from openpyxl.utils import get_column_letter


def analyze_excel_file(file_path: Path) -> dict[str, Any]:
    """Analyze an Excel file and extract its structure."""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        analysis = {"file_name": file_path.name, "sheets": []}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_info = {
                "name": sheet_name,
                "dimensions": sheet.dimensions,
                "max_row": sheet.max_row,
                "max_col": sheet.max_column,
                "headers": [],
                "sample_data": [],
                "empty_cells": [],
            }

            if sheet.max_row >= 1:
                for col in range(1, min(sheet.max_column + 1, 20)):
                    cell = sheet.cell(row=1, column=col)
                    if cell.value:
                        sheet_info["headers"].append({
                            "column": get_column_letter(col),
                            "value": str(cell.value)[:100],
                        })

            for row in range(2, min(sheet.max_row + 1, 7)):
                row_data = {}
                for col in range(1, min(sheet.max_column + 1, 20)):
                    cell = sheet.cell(row=row, column=col)
                    col_letter = get_column_letter(col)
                    if cell.value is not None:
                        row_data[col_letter] = str(cell.value)[:50]
                    else:
                        sheet_info["empty_cells"].append(f"{col_letter}{row}")
                if row_data:
                    sheet_info["sample_data"].append(row_data)

            empty_count = 0
            for row in range(2, min(sheet.max_row + 1, 50)):
                for col in range(1, min(sheet.max_column + 1, 20)):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value is None:
                        col_letter = get_column_letter(col)
                        if len(sheet_info["empty_cells"]) < 50:
                            if f"{col_letter}{row}" not in sheet_info["empty_cells"]:
                                sheet_info["empty_cells"].append(f"{col_letter}{row}")
                        empty_count += 1

            sheet_info["total_empty_cells"] = empty_count
            analysis["sheets"].append(sheet_info)

        workbook.close()
        return analysis

    except Exception as e:
        return {"error": str(e), "file_name": file_path.name}


def analyze_excel_file_full(file_path: Path) -> dict[str, Any]:
    """Analyze an Excel file and extract ALL row labels, headers, and cell values."""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        analysis = {"file_name": file_path.name, "sheets": []}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_col = min(sheet.max_column or 1, 30)
            max_row = min(sheet.max_row or 1, 200)

            headers = []
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=1, column=col)
                col_letter = get_column_letter(col)
                val = str(cell.value) if cell.value is not None else ""
                headers.append(f"{col_letter}: {val}")

            rows = []
            empty_cells = []

            for row_idx in range(2, max_row + 1):
                label_cell = sheet.cell(row=row_idx, column=1)
                label = str(label_cell.value) if label_cell.value is not None else ""
                has_any_data = label != ""
                cells = {}

                for col in range(2, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col)
                    col_letter = get_column_letter(col)
                    if cell.value is not None:
                        val = cell.value
                        if isinstance(val, float) and val == int(val):
                            cells[col_letter] = str(int(val))
                        else:
                            cells[col_letter] = str(val)
                        has_any_data = True
                    else:
                        cells[col_letter] = "EMPTY"
                        if label:
                            empty_cells.append(f"{col_letter}{row_idx}")

                if has_any_data:
                    rows.append({"row": row_idx, "label": label, "cells": cells})

            leftmost_date = None
            leftmost_period = None
            if sheet.max_column and sheet.max_column >= 2:
                b1 = sheet.cell(row=1, column=2).value
                b2 = sheet.cell(row=2, column=2).value
                leftmost_date = str(b1) if b1 is not None else None
                leftmost_period = str(b2) if b2 is not None else None

            data_rows = []
            for r in range(3, max_row + 1):
                if sheet.cell(row=r, column=2).value is not None:
                    data_rows.append(r)

            sheet_info = {
                "name": sheet_name,
                "max_row": max_row,
                "max_col": max_col,
                "headers": headers,
                "rows": rows,
                "empty_cells": empty_cells,
                "total_empty_cells": len(empty_cells),
                "leftmost_date": leftmost_date,
                "leftmost_period": leftmost_period,
                "data_rows": data_rows,
            }
            analysis["sheets"].append(sheet_info)

        workbook.close()
        return analysis

    except Exception as e:
        return {"error": str(e), "file_name": file_path.name}


def format_full_schema_for_llm(analysis: dict[str, Any]) -> str:
    """Format the full file analysis as a readable string for the LLM context."""
    if "error" in analysis:
        return f"Error analyzing file: {analysis['error']}"

    output = []
    output.append(f"File: {analysis['file_name']}")

    for sheet in analysis.get("sheets", []):
        output.append(f"\nSheet: {sheet['name']} ({sheet['max_row']} rows x {sheet['max_col']} cols)")
        output.append(f"Headers: {' | '.join(sheet['headers'])}")
        output.append("")

        for row_data in sheet["rows"]:
            row_num = row_data["row"]
            label = row_data["label"]
            cells_str = ", ".join(
                f"{col}={val}" for col, val in row_data["cells"].items()
            )
            output.append(f"  Row {row_num} [{label}]: {cells_str}")

        if sheet["empty_cells"]:
            output.append(f"\nEmpty cells needing data ({sheet['total_empty_cells']}): {', '.join(sheet['empty_cells'])}")

    return "\n".join(output)
