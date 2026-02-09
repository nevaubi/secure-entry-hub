"""
Excel file updater using openpyxl.

Handles updating specific cells in Excel files based on AI instructions.
"""

from copy import copy
from pathlib import Path
from typing import Any
import openpyxl
from openpyxl.utils import column_index_from_string


class ExcelUpdater:
    """Updates Excel files based on AI instructions."""

    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.changes_made = 0

    def update_cell(self, sheet_name: str, cell_ref: str, value: Any) -> bool:
        """
        Update a specific cell in the workbook.

        Args:
            sheet_name: Name of the sheet
            cell_ref: Cell reference like "A1" or "B5"
            value: Value to set

        Returns:
            True if successful, False otherwise
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                print(f"Sheet '{sheet_name}' not found")
                return False

            sheet = self.workbook[sheet_name]
            sheet[cell_ref] = value
            self.changes_made += 1
            print(f"Updated {sheet_name}!{cell_ref} = {value}")
            return True

        except Exception as e:
            print(f"Error updating cell {sheet_name}!{cell_ref}: {e}")
            return False

    def update_cells_batch(self, updates: list[dict]) -> int:
        """
        Update multiple cells at once.

        Args:
            updates: List of dicts with keys: sheet_name, cell_ref, value

        Returns:
            Number of successful updates
        """
        successful = 0
        for update in updates:
            if self.update_cell(
                update["sheet_name"],
                update["cell_ref"],
                update["value"]
            ):
                successful += 1
        return successful

    def save(self) -> bool:
        """
        Save the workbook.

        Returns:
            True if successful, False otherwise
        """
        try:
            self.workbook.save(self.file_path)
            print(f"Saved {self.file_path} with {self.changes_made} changes")
            return True
        except Exception as e:
            print(f"Error saving {self.file_path}: {e}")
            return False

    def close(self):
        """Close the workbook."""
        self.workbook.close()

    def insert_new_period_column(
        self, sheet_name: str, date_header: str, period_header: str
    ) -> dict:
        """
        Insert a new column B for a new fiscal period, shifting existing data right.

        Args:
            sheet_name: Name of the sheet
            date_header: Date for row 1, e.g. "2026-01-31"
            period_header: Period label for row 2, e.g. "Q4 2026"

        Returns:
            Dict with success status and list of row numbers needing data
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}

            sheet = self.workbook[sheet_name]

            # Insert a blank column at position 2 (column B), shifting everything right
            sheet.insert_cols(2)

            # Set headers in the new column B
            sheet.cell(row=1, column=2).value = date_header
            sheet.cell(row=2, column=2).value = period_header
            self.changes_made += 2

            # Copy header styling from column C (the shifted original) to new column B
            for row in [1, 2]:
                source_cell = sheet.cell(row=row, column=3)
                target_cell = sheet.cell(row=row, column=2)
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.border = copy(source_cell.border)
                target_cell.number_format = source_cell.number_format

            # Scan column C (the old column B, now shifted right) to find rows with data
            # Build a row_map with labels from column A for direct agent guidance
            data_rows = []
            row_map = []
            for row_idx in range(3, (sheet.max_row or 2) + 1):
                cell_c = sheet.cell(row=row_idx, column=3)
                if cell_c.value is not None:
                    data_rows.append(row_idx)
                    label = sheet.cell(row=row_idx, column=1).value or ""
                    row_map.append({"row": row_idx, "label": str(label).strip(), "cell": f"B{row_idx}"})

            print(f"Inserted new column B in {sheet_name}: {date_header} / {period_header}")
            print(f"  {len(data_rows)} rows need data (rows with values in adjacent column)")

            # Build a human-readable cell list for the agent
            cell_list = ", ".join(f"B{r} ({m['label']})" for r, m in zip(data_rows[:50], row_map[:50]))
            if len(data_rows) > 50:
                cell_list += f"... ({len(data_rows)} total)"

            return {
                "success": True,
                "data_rows": data_rows,
                "row_map": row_map,
                "total_rows_needing_data": len(data_rows),
                "message": f"New column B inserted with headers '{date_header}' / '{period_header}'. Fill these cells: {cell_list}",
            }

        except Exception as e:
            print(f"Error inserting column in {sheet_name}: {e}")
            return {"success": False, "error": str(e)}

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


def update_file(file_path: Path, updates: list[dict]) -> tuple[bool, int]:
    """
    Convenience function to update a file with multiple changes.

    Args:
        file_path: Path to the Excel file
        updates: List of cell updates

    Returns:
        Tuple of (success, number of changes)
    """
    with ExcelUpdater(file_path) as updater:
        count = updater.update_cells_batch(updates)
        if count > 0:
            success = updater.save()
            return success, count
        return True, 0
