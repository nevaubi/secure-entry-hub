"""
Excel schema analysis using Claude.

Dynamically understands the structure of Excel files to determine
what data needs to be updated.
"""

import os
from pathlib import Path
from typing import Any
import openpyxl
from openpyxl.utils import get_column_letter


def analyze_excel_file(file_path: Path) -> dict[str, Any]:
    """
    Analyze an Excel file and extract its structure.

    Args:
        file_path: Path to the Excel file

    Returns:
        Dict containing sheet names, headers, sample data, and empty cells
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        analysis = {
            "file_name": file_path.name,
            "sheets": [],
        }

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_info = {
                "name": sheet_name,
                "dimensions": sheet.dimensions,
                "max_row": sheet.max_row,
                "max_col": sheet.max_column,
                "headers": [],
                "sample_data": [],
                "empty_cells": [],
            }

            # Get headers (first row)
            if sheet.max_row >= 1:
                for col in range(1, min(sheet.max_column + 1, 20)):
                    cell = sheet.cell(row=1, column=col)
                    if cell.value:
                        sheet_info["headers"].append({
                            "column": get_column_letter(col),
                            "value": str(cell.value)[:100],
                        })

            # Get sample data (first 5 data rows)
            for row in range(2, min(sheet.max_row + 1, 7)):
                row_data = {}
                for col in range(1, min(sheet.max_column + 1, 20)):
                    cell = sheet.cell(row=row, column=col)
                    col_letter = get_column_letter(col)

                    if cell.value is not None:
                        row_data[col_letter] = str(cell.value)[:50]
                    else:
                        sheet_info["empty_cells"].append(f"{col_letter}{row}")

                if row_data:
                    sheet_info["sample_data"].append(row_data)

            # Look for cells that might need updating
            empty_count = 0
            for row in range(2, min(sheet.max_row + 1, 50)):
                for col in range(1, min(sheet.max_column + 1, 20)):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value is None:
                        col_letter = get_column_letter(col)
                        if len(sheet_info["empty_cells"]) < 50:
                            if f"{col_letter}{row}" not in sheet_info["empty_cells"]:
                                sheet_info["empty_cells"].append(f"{col_letter}{row}")
                        empty_count += 1

            sheet_info["total_empty_cells"] = empty_count
            analysis["sheets"].append(sheet_info)

        workbook.close()
        return analysis

    except Exception as e:
        return {"error": str(e), "file_name": file_path.name}


def analyze_excel_file_full(file_path: Path) -> dict[str, Any]:
    """
    Analyze an Excel file and extract ALL row labels, headers, and cell values.

    This provides a complete picture of the file contents so the agent can see
    exactly which cells are empty and what values already exist.

    Args:
        file_path: Path to the Excel file

    Returns:
        Dict with full grid data: headers, all rows with labels and cell values,
        and a list of empty cell references.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        analysis = {
            "file_name": file_path.name,
            "sheets": [],
        }

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_col = min(sheet.max_column or 1, 30)  # Cap at 30 columns
            max_row = min(sheet.max_row or 1, 200)     # Cap at 200 rows

            # Extract headers (row 1)
            headers = []
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=1, column=col)
                col_letter = get_column_letter(col)
                val = str(cell.value) if cell.value is not None else ""
                headers.append(f"{col_letter}: {val}")

            # Extract all data rows
            rows = []
            empty_cells = []

            for row_idx in range(2, max_row + 1):
                # Column A is the row label
                label_cell = sheet.cell(row=row_idx, column=1)
                label = str(label_cell.value) if label_cell.value is not None else ""

                # Skip completely empty rows (no label, no data)
                has_any_data = label != ""
                cells = {}

                for col in range(2, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col)
                    col_letter = get_column_letter(col)

                    if cell.value is not None:
                        # Convert to string, preserve full numeric precision
                        val = cell.value
                        if isinstance(val, float) and val == int(val):
                            cells[col_letter] = str(int(val))
                        else:
                            cells[col_letter] = str(val)
                        has_any_data = True
                    else:
                        cells[col_letter] = "EMPTY"
                        # Only track as empty if the row has a label
                        # (empty cells in label-less rows aren't meaningful)
                        if label:
                            empty_cells.append(f"{col_letter}{row_idx}")

                if has_any_data:
                    rows.append({
                        "row": row_idx,
                        "label": label,
                        "cells": cells,
                    })

            sheet_info = {
                "name": sheet_name,
                "max_row": max_row,
                "max_col": max_col,
                "headers": headers,
                "rows": rows,
                "empty_cells": empty_cells,
                "total_empty_cells": len(empty_cells),
            }

            analysis["sheets"].append(sheet_info)

        workbook.close()
        return analysis

    except Exception as e:
        return {"error": str(e), "file_name": file_path.name}


def format_full_schema_for_llm(analysis: dict[str, Any]) -> str:
    """
    Format the full file analysis as a readable string for the LLM context.

    Args:
        analysis: Result from analyze_excel_file_full()

    Returns:
        Formatted string with all rows, headers, and cell values
    """
    if "error" in analysis:
        return f"Error analyzing file: {analysis['error']}"

    output = []
    output.append(f"File: {analysis['file_name']}")

    for sheet in analysis.get("sheets", []):
        output.append(f"\nSheet: {sheet['name']} ({sheet['max_row']} rows x {sheet['max_col']} cols)")
        output.append(f"Headers: {' | '.join(sheet['headers'])}")
        output.append("")

        for row_data in sheet["rows"]:
            row_num = row_data["row"]
            label = row_data["label"]
            cells_str = ", ".join(
                f"{col}={val}" for col, val in row_data["cells"].items()
            )
            output.append(f"  Row {row_num} [{label}]: {cells_str}")

        if sheet["empty_cells"]:
            output.append(f"\nEmpty cells needing data ({sheet['total_empty_cells']}): {', '.join(sheet['empty_cells'])}")

    return "\n".join(output)


def analyze_all_files(files: dict[str, Path]) -> dict[str, dict]:
    """
    Analyze all Excel files for a ticker.

    Args:
        files: Dict mapping bucket names to local file paths

    Returns:
        Dict mapping bucket names to analysis results
    """
    analyses = {}

    for bucket_name, file_path in files.items():
        if file_path.exists():
            analyses[bucket_name] = analyze_excel_file(file_path)
        else:
            analyses[bucket_name] = {"error": "File not found", "file_name": file_path.name}

    return analyses


def format_schema_for_llm(analyses: dict[str, dict]) -> str:
    """
    Format the schema analysis as a string for the LLM context.

    Args:
        analyses: Dict of analysis results per file

    Returns:
        Formatted string describing all file schemas
    """
    output = []

    for bucket_name, analysis in analyses.items():
        output.append(f"\n## File: {bucket_name}")

        if "error" in analysis:
            output.append(f"  Error: {analysis['error']}")
            continue

        for sheet in analysis.get("sheets", []):
            output.append(f"\n  ### Sheet: {sheet['name']}")
            output.append(f"  Dimensions: {sheet['dimensions']} ({sheet['max_row']} rows x {sheet['max_col']} cols)")

            if sheet["headers"]:
                header_str = ", ".join([f"{h['column']}: {h['value']}" for h in sheet["headers"]])
                output.append(f"  Headers: {header_str}")

            if sheet["sample_data"]:
                output.append("  Sample data:")
                for i, row in enumerate(sheet["sample_data"][:3], start=2):
                    row_str = ", ".join([f"{k}: {v}" for k, v in row.items()])
                    output.append(f"    Row {i}: {row_str}")

            if sheet["empty_cells"]:
                output.append(f"  Empty cells (sample): {', '.join(sheet['empty_cells'][:20])}")
                if sheet["total_empty_cells"] > 20:
                    output.append(f"  (Total empty cells: {sheet['total_empty_cells']})")

    return "\n".join(output)
